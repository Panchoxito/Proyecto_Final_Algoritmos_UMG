import openpyxl

def guardar_resultado(nombre_archivo, hoja, datos):
    """Guarda los resultados en un archivo Excel."""
    try:
        libro = openpyxl.load_workbook(nombre_archivo)
    except FileNotFoundError:
        libro = openpyxl.Workbook()

    if hoja not in libro.sheetnames:
        libro.create_sheet(hoja)

    ws = libro[hoja]
    ws.append(datos)
    libro.save(nombre_archivo)
