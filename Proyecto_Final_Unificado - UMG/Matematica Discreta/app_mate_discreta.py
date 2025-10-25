# app_mate_discreta.py
# -*- coding: utf-8 -*-

import os
import datetime
import tkinter as tk
from tkinter import ttk, messagebox
from math import factorial
from openpyxl import Workbook, load_workbook

# ---------------------------
# Paleta de colores (pastel)
# ---------------------------
PALETA = {
    "fondo":      "#F7F7FA",
    "amarillo":   "#FFF7AE",
    "menta":      "#BFFCC6",
    "durazno":    "#FFCCB6",
    "aqua":       "#A0E7E5",
    "rosa":       "#FFC6FF",
    "lavanda":    "#BDB2FF",
    "texto":      "#323232",
    "titulo":     "#2B2B2B",
    "borde":      "#E3E3EE",
}

ARCHIVO_EXCEL = "datos_mate_discreta.xlsx"

# ---------------------------
# Utilidades: Excel
# ---------------------------
def asegurar_excel():
    """Crea el archivo y hojas si no existen."""
    if not os.path.exists(ARCHIVO_EXCEL):
        wb = Workbook()
        # Hoja 1: combinatoria
        ws1 = wb.active
        ws1.title = "combinatoria"
        ws1.append(["fecha_hora", "operacion", "n", "r", "resultado"])
        # Hoja 2: conjuntos
        ws2 = wb.create_sheet("conjuntos")
        ws2.append(["fecha_hora", "operacion", "conjunto_A", "conjunto_B", "resultado"])
        # Hoja 3: euclides
        ws3 = wb.create_sheet("euclides")
        ws3.append(["fecha_hora", "a", "b", "mcd", "pasos"])
        wb.save(ARCHIVO_EXCEL)

def guardar_combinatoria(operacion, n, r, resultado):
    asegurar_excel()
    wb = load_workbook(ARCHIVO_EXCEL)
    ws = wb["combinatoria"]
    ws.append([datetime.datetime.now(), operacion, n, r, resultado])
    wb.save(ARCHIVO_EXCEL)

def guardar_conjuntos(operacion, a_str, b_str, resultado_str):
    asegurar_excel()
    wb = load_workbook(ARCHIVO_EXCEL)
    ws = wb["conjuntos"]
    ws.append([datetime.datetime.now(), operacion, a_str, b_str, resultado_str])
    wb.save(ARCHIVO_EXCEL)

def guardar_euclides(a, b, mcd, pasos):
    asegurar_excel()
    wb = load_workbook(ARCHIVO_EXCEL)
    ws = wb["euclides"]
    ws.append([datetime.datetime.now(), a, b, mcd, " | ".join(pasos)])
    wb.save(ARCHIVO_EXCEL)

# ---------------------------
# Utilidades: lógica
# ---------------------------
def combinaciones(n, r):
    if r < 0 or r > n:
        raise ValueError("r debe estar entre 0 y n")
    return factorial(n) // (factorial(r) * factorial(n - r))

def permutaciones(n, r):
    if r < 0 or r > n:
        raise ValueError("r debe estar entre 0 y n")
    return factorial(n) // factorial(n - r)

def parsear_conjunto(texto):
    """
    Convierte '1, 2, 3, a' -> conjunto de cadenas {'1','2','3','a'}
    (mantenemos strings para permitir símbolos).
    """
    if not texto.strip():
        return set()
    return set(s.strip() for s in texto.split(",") if s.strip() != "")

def formatear_conjunto(cjto):
    return "{" + ", ".join(sorted(cjto, key=lambda x: str(x))) + "}"

def mcd_euclides(a, b):
    a, b = abs(int(a)), abs(int(b))
    pasos = []
    while b != 0:
        q = a // b
        r = a % b
        pasos.append(f"{a} = {b} × {q} + {r}")
        a, b = b, r
    pasos.append(f"MCD = {a}")
    return a, pasos

# ---------------------------
# UI: componentes reutilizables
# ---------------------------
class Tarjeta(ttk.Frame):
    """Contenedor con borde suave."""
    def __init__(self, master, color_fondo, **kwargs):
        super().__init__(master, **kwargs)
        self.configure(style="Tarjeta.TFrame")
        self.interior = tk.Frame(self, bg=color_fondo, bd=0, highlightthickness=0)
        self.interior.pack(fill="both", expand=True, padx=2, pady=2)
        self.interior.grid_columnconfigure(0, weight=1)

class BotonPastel(tk.Button):
    def __init__(self, master, texto, color, comando):
        super().__init__(
            master,
            text=texto,
            command=comando,
            bg=color,
            activebackground=color,
            fg=PALETA["texto"],
            relief="flat",
            bd=0,
            padx=22,
            pady=14,
            cursor="hand2",
        )
        self.configure(font=("Segoe UI", 12, "bold"))
        self.bind("<Enter>", lambda e: self.configure(relief="groove"))
        self.bind("<Leave>", lambda e: self.configure(relief="flat"))

# ---------------------------
# App principal
# ---------------------------
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Proyecto Matemática Discreta")
        self.geometry("920x680")
        self.minsize(860, 620)
        self.configure(bg=PALETA["fondo"])

        estilo = ttk.Style(self)
        estilo.theme_use("clam")
        estilo.configure("Tarjeta.TFrame", background=PALETA["borde"], relief="flat")
        estilo.configure("TLabel", background=PALETA["fondo"], foreground=PALETA["texto"])
        estilo.configure("Titulo.TLabel", font=("Segoe UI", 22, "bold"), foreground=PALETA["titulo"])

        self.columnconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)

        self.lbl_titulo = ttk.Label(self, text="Proyecto: Matemática Discreta", style="Titulo.TLabel")
        self.lbl_titulo.grid(row=0, column=0, pady=(24, 8))

        self.contenedor = ttk.Frame(self)
        self.contenedor.grid(row=1, column=0, sticky="nsew", padx=24, pady=16)
        self.contenedor.columnconfigure(0, weight=1)
        self.contenedor.rowconfigure(0, weight=1)

        self.pantallas = {}
        self.mostrar_pantalla(self.pantalla_inicio())

    # ---------- pantallas ----------
    def pantalla_inicio(self):
        if "inicio" in self.pantallas:
            return self.pantallas["inicio"]

        marco = ttk.Frame(self.contenedor)
        marco.grid_columnconfigure(0, weight=1)

        tarjeta = Tarjeta(marco, PALETA["aqua"])
        tarjeta.grid(row=0, column=0, padx=180, pady=80, sticky="n")
        lbl = tk.Label(
            tarjeta.interior,
            text="Matemática Discreta",
            bg=PALETA["aqua"],
            fg=PALETA["texto"],
            font=("Segoe UI", 20, "bold"),
        )
        lbl.grid(row=0, column=0, pady=(24, 12), sticky="n")

        BotonPastel(
            tarjeta.interior, "Entrar", PALETA["lavanda"],
            lambda: self.mostrar_pantalla(self.pantalla_menu_discreta())
        ).grid(row=1, column=0, pady=(8, 24), sticky="ew")

        self.pantallas["inicio"] = marco
        return marco

    def pantalla_menu_discreta(self):
        if "menu" in self.pantallas:
            return self.pantallas["menu"]

        marco = ttk.Frame(self.contenedor)
        marco.grid_columnconfigure(0, weight=1)

        subt = ttk.Label(marco, text="Selecciona un módulo", font=("Segoe UI", 14))
        subt.grid(row=0, column=0, pady=(8, 16))

        # Tarjeta con botones VERTICALES centrados
        tarjeta = Tarjeta(marco, PALETA["fondo"])
        tarjeta.grid(row=1, column=0, padx=280, pady=10, sticky="n")
        t = tarjeta.interior
        t.grid_columnconfigure(0, weight=1)

        BotonPastel(
            t, "Combinaciones y Permutaciones", PALETA["amarillo"],
            lambda: self.mostrar_pantalla(self.pantalla_combinatoria())
        ).grid(row=0, column=0, padx=12, pady=8, sticky="ew")

        BotonPastel(
            t, "Conjuntos", PALETA["menta"],
            lambda: self.mostrar_pantalla(self.pantalla_conjuntos())
        ).grid(row=1, column=0, padx=12, pady=8, sticky="ew")

        BotonPastel(
            t, "Euclides (MCD)", PALETA["durazno"],
            lambda: self.mostrar_pantalla(self.pantalla_euclides())
        ).grid(row=2, column=0, padx=12, pady=8, sticky="ew")

        BotonPastel(
            marco, "Volver", PALETA["rosa"],
            lambda: self.mostrar_pantalla(self.pantalla_inicio())
        ).grid(row=2, column=0, pady=(12, 0))

        self.pantallas["menu"] = marco
        return marco

    def pantalla_combinatoria(self):
        if "combinatoria" in self.pantallas:
            return self.pantallas["combinatoria"]

        marco = ttk.Frame(self.contenedor)
        marco.grid_columnconfigure(0, weight=1)

        ttk.Label(marco, text="Combinaciones y Permutaciones", style="Titulo.TLabel").grid(row=0, column=0, pady=(0, 12))

        tarjeta = Tarjeta(marco, PALETA["fondo"])
        tarjeta.grid(row=1, column=0, padx=120, pady=10, sticky="n")
        f = tarjeta.interior
        f.grid_columnconfigure(1, weight=1)

        ttk.Label(f, text="Operación:").grid(row=0, column=0, sticky="e", pady=6, padx=6)
        combo_operacion = ttk.Combobox(
            f, state="readonly",
            values=["Todas", "Combinaciones (nCr)", "Permutaciones (nPr)"]
        )
        combo_operacion.current(0)
        combo_operacion.grid(row=0, column=1, sticky="we", pady=6, padx=6)

        ttk.Label(f, text="n:").grid(row=1, column=0, sticky="e", padx=6, pady=6)
        entrada_n = ttk.Entry(f)
        entrada_n.grid(row=1, column=1, sticky="we", padx=6, pady=6)

        ttk.Label(f, text="r:").grid(row=2, column=0, sticky="e", padx=6, pady=6)
        entrada_r = ttk.Entry(f)
        entrada_r.grid(row=2, column=1, sticky="we", padx=6, pady=6)

        salida = tk.Text(f, height=7, bg="white", relief="flat", bd=1, highlightthickness=1,
                         highlightbackground=PALETA["borde"])
        salida.grid(row=3, column=0, columnspan=2, sticky="we", padx=6, pady=(10, 6))

        def calcular():
            try:
                n = int(entrada_n.get())
                r = int(entrada_r.get())
                op = combo_operacion.get()
                salida.delete("1.0", "end")
                if op == "Todas":
                    res_c = combinaciones(n, r)
                    res_p = permutaciones(n, r)
                    salida.insert("end", f"C({n}, {r}) = {res_c}\n")
                    salida.insert("end", f"P({n}, {r}) = {res_p}\n")
                    guardar_combinatoria("combinaciones", n, r, res_c)
                    guardar_combinatoria("permutaciones", n, r, res_p)
                elif op.startswith("Comb"):
                    res = combinaciones(n, r)
                    salida.insert("end", f"C({n}, {r}) = {res}\n")
                    guardar_combinatoria("combinaciones", n, r, res)
                else:
                    res = permutaciones(n, r)
                    salida.insert("end", f"P({n}, {r}) = {res}\n")
                    guardar_combinatoria("permutaciones", n, r, res)
            except ValueError as e:
                messagebox.showerror("Error", f"Entrada inválida: {e}")

        BotonPastel(f, "Calcular", PALETA["lavanda"], calcular).grid(row=4, column=0, columnspan=2, padx=6, pady=8, sticky="ew")

        BotonPastel(marco, "Volver", PALETA["rosa"], lambda: self.mostrar_pantalla(self.pantalla_menu_discreta())).grid(row=2, column=0, pady=8)

        self.pantallas["combinatoria"] = marco
        return marco

    def pantalla_conjuntos(self):
        if "conjuntos" in self.pantallas:
            return self.pantallas["conjuntos"]

        marco = ttk.Frame(self.contenedor)
        marco.grid_columnconfigure(0, weight=1)

        ttk.Label(marco, text="Operaciones con Conjuntos", style="Titulo.TLabel").grid(row=0, column=0, pady=(0, 12))

        tarjeta = Tarjeta(marco, PALETA["fondo"])
        tarjeta.grid(row=1, column=0, padx=120, pady=10, sticky="n")
        f = tarjeta.interior
        f.grid_columnconfigure(1, weight=1)

        ttk.Label(f, text="Operación:").grid(row=0, column=0, sticky="e", padx=6, pady=6)
        combo_op = ttk.Combobox(
            f, state="readonly",
            values=[
                "Todas",
                "Unión (A ∪ B)",
                "Intersección (A ∩ B)",
                "Diferencia (A − B)",
                "Diferencia (B − A)",
                "Diferencia simétrica (A △ B)",
                "Subconjunto (A ⊆ B)",
                "Subconjunto (B ⊆ A)"
            ]
        )
        combo_op.current(0)
        combo_op.grid(row=0, column=1, sticky="we", padx=6, pady=6)

        ttk.Label(f, text="Conjunto A (separa por comas):").grid(row=1, column=0, sticky="e", padx=6, pady=6)
        entrada_a = ttk.Entry(f)
        entrada_a.grid(row=1, column=1, sticky="we", padx=6, pady=6)

        ttk.Label(f, text="Conjunto B (separa por comas):").grid(row=2, column=0, sticky="e", padx=6, pady=6)
        entrada_b = ttk.Entry(f)
        entrada_b.grid(row=2, column=1, sticky="we", padx=6, pady=6)

        salida = tk.Text(f, height=10, bg="white", relief="flat", bd=1, highlightthickness=1,
                         highlightbackground=PALETA["borde"])
        salida.grid(row=3, column=0, columnspan=2, sticky="we", padx=6, pady=(10, 6))

        def ejecutar():
            op = combo_op.get()
            a_str = entrada_a.get()
            b_str = entrada_b.get()
            A = parsear_conjunto(a_str)
            B = parsear_conjunto(b_str)

            salida.delete("1.0", "end")

            def _guardar_y_mostrar(nombre, cjto):
                texto = f"{nombre} = {formatear_conjunto(cjto)}"
                salida.insert("end", texto + "\n")
                guardar_conjuntos(nombre.lower(), a_str, b_str, formatear_conjunto(cjto))

            if op == "Todas":
                _guardar_y_mostrar("A ∪ B", A.union(B))
                _guardar_y_mostrar("A ∩ B", A.intersection(B))
                _guardar_y_mostrar("A − B", A.difference(B))
                _guardar_y_mostrar("B − A", B.difference(A))
                _guardar_y_mostrar("A △ B", A.symmetric_difference(B))
                es_sub_a = A.issubset(B)
                texto_a = f"A ⊆ B  →  {es_sub_a}"
                salida.insert("end", texto_a + "\n")
                guardar_conjuntos("subconjunto", a_str, b_str, texto_a)
                es_sub_b = B.issubset(A)
                texto_b = f"B ⊆ A  →  {es_sub_b}"
                salida.insert("end", texto_b + "\n")
                guardar_conjuntos("subconjunto", a_str, b_str, texto_b)
            elif op.startswith("Unión (A ∪ B)"):
                _guardar_y_mostrar("A ∪ B", A.union(B))
            elif op.startswith("Intersección (A ∩ B)"):
                _guardar_y_mostrar("A ∩ B", A.intersection(B))
            elif op.startswith("Diferencia (A − B)"):
                _guardar_y_mostrar("A − B", A.difference(B))
            elif op.startswith("Diferencia (B − A)"):
                _guardar_y_mostrar("B − A", B.difference(A))
            elif op.startswith("Diferencia simétrica (A △ B)"):
                _guardar_y_mostrar("A △ B", A.symmetric_difference(B))
            elif op.startswith("Subconjunto (A ⊆ B)"):
                es_sub = A.issubset(B)
                texto = f"A ⊆ B  →  {es_sub}"
                salida.insert("end", texto + "\n")
                guardar_conjuntos("subconjunto", a_str, b_str, texto)
            elif op.startswith("Subconjunto (B ⊆ A)"):
                es_sub = B.issubset(A)
                texto = f"B ⊆ A  →  {es_sub}"
                salida.insert("end", texto + "\n")
                guardar_conjuntos("subconjunto", a_str, b_str, texto)

        BotonPastel(f, "Ejecutar", PALETA["lavanda"], ejecutar).grid(row=4, column=0, columnspan=2, padx=6, pady=8, sticky="ew")

        BotonPastel(marco, "Volver", PALETA["rosa"], lambda: self.mostrar_pantalla(self.pantalla_menu_discreta())).grid(row=2, column=0, pady=8)

        self.pantallas["conjuntos"] = marco
        return marco

    def pantalla_euclides(self):
        if "euclides" in self.pantallas:
            return self.pantallas["euclides"]

        marco = ttk.Frame(self.contenedor)
        marco.grid_columnconfigure(0, weight=1)

        ttk.Label(marco, text="Algoritmo de Euclides (MCD)", style="Titulo.TLabel").grid(row=0, column=0, pady=(0, 12))

        tarjeta = Tarjeta(marco, PALETA["fondo"])
        tarjeta.grid(row=1, column=0, padx=120, pady=10, sticky="n")
        f = tarjeta.interior
        f.grid_columnconfigure(1, weight=1)

        ttk.Label(f, text="Operación:").grid(row=0, column=0, sticky="e", padx=6, pady=6)
        combo_op = ttk.Combobox(
            f, state="readonly",
            values=["Calcular MCD (Euclides)", "Todas"]
        )
        combo_op.current(1)  # Por defecto "Todas"
        combo_op.grid(row=0, column=1, sticky="we", padx=6, pady=6)

        ttk.Label(f, text="a:").grid(row=1, column=0, sticky="e", padx=6, pady=6)
        entrada_a = ttk.Entry(f)
        entrada_a.grid(row=1, column=1, sticky="we", padx=6, pady=6)

        ttk.Label(f, text="b:").grid(row=2, column=0, sticky="e", padx=6, pady=6)
        entrada_b = ttk.Entry(f)
        entrada_b.grid(row=2, column=1, sticky="we", padx=6, pady=6)

        salida = tk.Text(f, height=10, bg="white", relief="flat", bd=1, highlightthickness=1,
                         highlightbackground=PALETA["borde"])
        salida.grid(row=3, column=0, columnspan=2, sticky="we", padx=6, pady=(10, 6))

        def calcular():
            try:
                a = int(entrada_a.get())
                b = int(entrada_b.get())
                # Para "Todas" o "Calcular", el procedimiento es el mismo: mostrar todos los pasos.
                m, pasos = mcd_euclides(a, b)
                salida.delete("1.0", "end")
                for p in pasos:
                    salida.insert("end", p + "\n")
                guardar_euclides(a, b, m, pasos)
            except ValueError:
                messagebox.showerror("Error", "Ingresa enteros válidos para a y b.")

        BotonPastel(f, "Calcular", PALETA["lavanda"], calcular).grid(row=4, column=0, columnspan=2, padx=6, pady=8, sticky="ew")

        BotonPastel(marco, "Volver", PALETA["rosa"], lambda: self.mostrar_pantalla(self.pantalla_menu_discreta())).grid(row=2, column=0, pady=8)

        self.pantallas["euclides"] = marco
        return marco

    # ---------- util ----------
    def mostrar_pantalla(self, marco):
        # Ocultar todas y mostrar la solicitada
        for w in self.contenedor.winfo_children():
            w.grid_forget()
        marco.grid(row=0, column=0, sticky="n")

# ---------------------------
# Main
# ---------------------------
if __name__ == "__main__":
    asegurar_excel()
    app = App()
    app.mainloop()
